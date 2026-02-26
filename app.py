from flask import (Flask, render_template, request, redirect,
                   url_for, session, flash, jsonify, send_file)
import sqlite3, os, io, json
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = 'kavisDairyMilkSecret2026!'
DATABASE = 'dairy.db'

# ── DB helpers ────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    db = get_db()
    db.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            name      TEXT    NOT NULL,
            username  TEXT    UNIQUE NOT NULL,
            email     TEXT,
            phone     TEXT,
            password  TEXT    NOT NULL,
            is_admin  INTEGER NOT NULL DEFAULT 0,
            created_at TEXT   DEFAULT (datetime('now'))
        );
        CREATE TABLE IF NOT EXISTS products (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            name        TEXT    NOT NULL,
            description TEXT,
            price       REAL    NOT NULL,
            unit        TEXT    DEFAULT 'litre',
            emoji       TEXT    DEFAULT '🥛',
            stock       INTEGER DEFAULT 100,
            category    TEXT    DEFAULT 'Milk'
        );
        CREATE TABLE IF NOT EXISTS orders (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id     INTEGER NOT NULL,
            total       REAL    NOT NULL,
            status      TEXT    DEFAULT 'Confirmed',
            address     TEXT,
            created_at  TEXT    DEFAULT (datetime('now')),
            FOREIGN KEY (user_id) REFERENCES users(id)
        );
        CREATE TABLE IF NOT EXISTS order_items (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id    INTEGER NOT NULL,
            product_id  INTEGER NOT NULL,
            product_name TEXT   NOT NULL,
            quantity    INTEGER NOT NULL,
            price       REAL    NOT NULL,
            FOREIGN KEY (order_id)    REFERENCES orders(id),
            FOREIGN KEY (product_id)  REFERENCES products(id)
        );
    ''')

    # Seed admin
    if not db.execute("SELECT id FROM users WHERE username='admin'").fetchone():
        db.execute("INSERT INTO users (name,username,email,password,is_admin) VALUES (?,?,?,?,?)",
                   ('Administrator', 'admin', 'admin@kavifarm.com',
                    generate_password_hash('admin123'), 1))

    # Seed products
    if not db.execute("SELECT id FROM products LIMIT 1").fetchone():
        products = [
            ('Full Cream Milk',    'Pure whole milk, rich in calcium and vitamins', 55,  'litre',  '🥛', 200, 'Milk'),
            ('Toned Milk',         'Low fat toned milk, ideal for daily use',        45,  'litre',  '🥛', 180, 'Milk'),
            ('Double Toned Milk',  'Very low fat double toned milk',                  38,  'litre',  '🥛', 150, 'Milk'),
            ('Organic A2 Milk',    'Premium A2 desi cow milk, highest quality',      90,  '500 ml', '🥛', 100, 'Milk'),
            ('Fresh Butter',       'Creamy churned butter made every morning',       120, '500 g',  '🧈', 80,  'Butter'),
            ('Desi Ghee',          'Pure clarified butter from A2 cow milk',         550, '500 ml', '🫙', 60,  'Ghee'),
            ('Natural Yoghurt',    'Thick, probiotic-rich homemade curd',            65,  '500 g',  '🍶', 120, 'Curd'),
            ('Mishti Doi',         'Sweetened Bengali-style yoghurt dessert',        80,  '250 g',  '🍮', 90,  'Curd'),
            ('Fresh Paneer',       'Soft cottage cheese made fresh daily',           180, '250 g',  '🧀', 70,  'Paneer'),
            ('Smoked Paneer',      'Lightly smoked artisan paneer block',            220, '250 g',  '🧀', 40,  'Paneer'),
            ('Buttermilk (Chaas)', 'Spiced refreshing buttermilk',                   25,  '500 ml', '🥤', 150, 'Drinks'),
            ('Mango Lassi',        'Sweet mango yoghurt drink',                      50,  '300 ml', '🥭', 100, 'Drinks'),
            ('Malai Cream',        'Rich fresh cream, perfect for cooking',          90,  '200 ml', '🍦', 80,  'Cream'),
            ('Milk Khoa (Mawa)',   'Reduced milk solid for sweets',                  280, '250 g',  '🍬', 50,  'Sweets'),
        ]
        db.executemany(
            "INSERT INTO products (name,description,price,unit,emoji,stock,category) VALUES (?,?,?,?,?,?,?)",
            products)

    db.commit()
    db.close()

# ── Context processor ─────────────────────────────────────────────────────────

@app.context_processor
def inject_cart_count():
    cart = session.get('cart', {})
    total = sum(item['qty'] for item in cart.values())
    return {'cart_count': total}

# ── Pages ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    db = get_db()
    products = db.execute("SELECT * FROM products ORDER BY RANDOM() LIMIT 6").fetchall()
    db.close()
    return render_template('index.html', products=products)

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/products')
def products():
    db = get_db()
    category = request.args.get('category', 'All')
    if category == 'All':
        prods = db.execute("SELECT * FROM products ORDER BY category, name").fetchall()
    else:
        prods = db.execute("SELECT * FROM products WHERE category=? ORDER BY name", (category,)).fetchall()
    categories = [r['category'] for r in db.execute("SELECT DISTINCT category FROM products ORDER BY category").fetchall()]
    db.close()
    return render_template('products.html', products=prods, categories=categories, active_cat=category)

@app.route('/contact')
def contact():
    return render_template('contact.html')

# ── Auth ──────────────────────────────────────────────────────────────────────

@app.route('/register', methods=['GET','POST'])
def register():
    if request.method == 'POST':
        name     = request.form['name'].strip()
        username = request.form['username'].strip()
        email    = request.form['email'].strip()
        phone    = request.form['phone'].strip()
        password = request.form['password']
        confirm  = request.form['confirm']

        if len(password) < 6:
            flash('Password must be at least 6 characters.', 'danger')
            return render_template('register.html')
        if password != confirm:
            flash('Passwords do not match.', 'danger')
            return render_template('register.html')

        db = get_db()
        try:
            db.execute("INSERT INTO users (name,username,email,phone,password) VALUES (?,?,?,?,?)",
                       (name, username, email, phone, generate_password_hash(password)))
            db.commit()

            # Save to Excel
            _append_user_to_excel(name, username, email, phone)

            flash('Registration successful! Please login.', 'success')
            return redirect(url_for('login'))
        except sqlite3.IntegrityError:
            flash('Username already exists. Choose another.', 'danger')
        finally:
            db.close()
    return render_template('register.html')

@app.route('/login', methods=['GET','POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']
        db   = get_db()
        user = db.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
        db.close()
        if user and check_password_hash(user['password'], password):
            session['user_id']  = user['id']
            session['username'] = user['username']
            session['name']     = user['name']
            session['is_admin'] = bool(user['is_admin'])
            flash(f"Welcome back, {user['name']}! 👋", 'success')
            return redirect(url_for('dashboard'))
        flash('Invalid username or password.', 'danger')
    return render_template('login.html')

@app.route('/admin-login', methods=['GET','POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password']
        db   = get_db()
        user = db.execute("SELECT * FROM users WHERE username=? AND is_admin=1", (username,)).fetchone()
        db.close()
        if user and check_password_hash(user['password'], password):
            session['user_id']  = user['id']
            session['username'] = user['username']
            session['name']     = user['name']
            session['is_admin'] = True
            flash('Admin access granted! 🛡️', 'success')
            return redirect(url_for('dashboard'))
        flash('Invalid admin credentials.', 'danger')
    return render_template('admin_login.html')

@app.route('/logout')
def logout():
    name = session.get('name', 'User')
    session.clear()
    flash(f'Goodbye, {name}! See you soon. 👋', 'info')
    return redirect(url_for('index'))

# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        flash('Please login first.', 'info')
        return redirect(url_for('login'))
    db = get_db()
    if session.get('is_admin'):
        users         = db.execute("SELECT * FROM users ORDER BY id DESC").fetchall()
        all_orders    = db.execute("""
            SELECT o.*, u.name, u.username FROM orders o
            JOIN users u ON o.user_id=u.id ORDER BY o.id DESC""").fetchall()
        products      = db.execute("SELECT * FROM products ORDER BY category, name").fetchall()
        total_revenue = db.execute("SELECT SUM(total) FROM orders").fetchone()[0] or 0
        db.close()
        return render_template('dashboard.html', users=users, all_orders=all_orders,
                               products=products, total_revenue=total_revenue)
    else:
        orders = db.execute("""
            SELECT * FROM orders WHERE user_id=? ORDER BY id DESC""", (session['user_id'],)).fetchall()
        db.close()
        return render_template('dashboard.html', orders=orders)

# ── Admin CRUD for products ───────────────────────────────────────────────────

@app.route('/admin/product/add', methods=['POST'])
def add_product():
    if not session.get('is_admin'): return redirect(url_for('index'))
    db = get_db()
    db.execute("INSERT INTO products (name,description,price,unit,emoji,stock,category) VALUES (?,?,?,?,?,?,?)",
               (request.form['name'], request.form['description'], float(request.form['price']),
                request.form['unit'], request.form['emoji'], int(request.form['stock']),
                request.form['category']))
    db.commit(); db.close()
    flash('Product added!', 'success')
    return redirect(url_for('dashboard'))

@app.route('/admin/product/edit/<int:pid>', methods=['POST'])
def edit_product(pid):
    if not session.get('is_admin'): return redirect(url_for('index'))
    db = get_db()
    db.execute("UPDATE products SET name=?,description=?,price=?,unit=?,emoji=?,stock=?,category=? WHERE id=?",
               (request.form['name'], request.form['description'], float(request.form['price']),
                request.form['unit'], request.form['emoji'], int(request.form['stock']),
                request.form['category'], pid))
    db.commit(); db.close()
    flash('Product updated!', 'success')
    return redirect(url_for('dashboard'))

@app.route('/admin/product/delete/<int:pid>', methods=['POST'])
def delete_product(pid):
    if not session.get('is_admin'): return redirect(url_for('index'))
    db = get_db()
    db.execute("DELETE FROM products WHERE id=?", (pid,))
    db.commit(); db.close()
    flash('Product deleted.', 'success')
    return redirect(url_for('dashboard'))

# ── Shop / Cart ───────────────────────────────────────────────────────────────

@app.route('/shop')
def shop():
    if 'user_id' not in session:
        flash('Please login to shop.', 'info')
        return redirect(url_for('login'))
    db = get_db()
    prods = db.execute("SELECT * FROM products ORDER BY category, name").fetchall()
    db.close()
    cart = session.get('cart', {})
    return render_template('shop.html', products=prods, cart=cart)

@app.route('/cart/add', methods=['POST'])
def cart_add():
    if 'user_id' not in session:
        return jsonify({'error': 'not_logged_in'}), 401
    pid   = str(request.form.get('product_id'))
    qty   = int(request.form.get('qty', 1))
    db    = get_db()
    prod  = db.execute("SELECT * FROM products WHERE id=?", (pid,)).fetchone()
    db.close()
    if not prod:
        return jsonify({'error': 'not_found'}), 404
    cart  = session.get('cart', {})
    if pid in cart:
        cart[pid]['qty'] += qty
    else:
        cart[pid] = {'id': pid, 'name': prod['name'], 'price': prod['price'],
                     'unit': prod['unit'], 'emoji': prod['emoji'], 'qty': qty}
    session['cart'] = cart
    total_items = sum(i['qty'] for i in cart.values())
    return jsonify({'success': True, 'cart_count': total_items})

@app.route('/cart/remove', methods=['POST'])
def cart_remove():
    pid  = str(request.form.get('product_id'))
    cart = session.get('cart', {})
    cart.pop(pid, None)
    session['cart'] = cart
    return jsonify({'success': True})

@app.route('/cart/update', methods=['POST'])
def cart_update():
    pid  = str(request.form.get('product_id'))
    qty  = int(request.form.get('qty', 1))
    cart = session.get('cart', {})
    if pid in cart:
        if qty <= 0:
            cart.pop(pid)
        else:
            cart[pid]['qty'] = qty
    session['cart'] = cart
    total = sum(i['price'] * i['qty'] for i in cart.values())
    return jsonify({'success': True, 'total': total})

@app.route('/checkout', methods=['GET','POST'])
def checkout():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    cart = session.get('cart', {})
    if not cart:
        flash('Your cart is empty!', 'info')
        return redirect(url_for('shop'))
    if request.method == 'POST':
        address = request.form.get('address', '').strip()
        total   = sum(i['price'] * i['qty'] for i in cart.values())
        db      = get_db()
        cur     = db.execute("INSERT INTO orders (user_id,total,address) VALUES (?,?,?)",
                             (session['user_id'], total, address))
        oid     = cur.lastrowid
        for item in cart.values():
            db.execute("INSERT INTO order_items (order_id,product_id,product_name,quantity,price) VALUES (?,?,?,?,?)",
                       (oid, item['id'], item['name'], item['qty'], item['price']))
            db.execute("UPDATE products SET stock=MAX(0,stock-?) WHERE id=?", (item['qty'], item['id']))
        db.commit()

        # Save order to Excel
        _append_order_to_excel(oid, session['name'], session['username'], address, total, cart)

        session.pop('cart', None)
        db.close()
        flash('Order placed successfully! 🎉', 'success')
        return redirect(url_for('billing', order_id=oid))
    items  = list(cart.values())
    subtotal = sum(i['price'] * i['qty'] for i in items)
    gst    = round(subtotal * 0.05, 2)
    total  = round(subtotal + gst, 2)
    return render_template('checkout.html', items=items, subtotal=subtotal, gst=gst, total=total)

@app.route('/billing/<int:order_id>')
def billing(order_id):
    if 'user_id' not in session:
        return redirect(url_for('login'))
    db    = get_db()
    order = db.execute("""
        SELECT o.*, u.name, u.username, u.phone, u.email
        FROM orders o JOIN users u ON o.user_id=u.id WHERE o.id=?""", (order_id,)).fetchone()
    items = db.execute("SELECT * FROM order_items WHERE order_id=?", (order_id,)).fetchall()
    db.close()
    if not order:
        flash('Order not found.', 'danger')
        return redirect(url_for('dashboard'))
    gst      = round(order['total'] * 0.05 / 1.05, 2)
    subtotal = round(order['total'] - gst, 2)
    return render_template('billing.html', order=order, items=items, gst=gst, subtotal=subtotal)

# ── Excel export ──────────────────────────────────────────────────────────────

EXCEL_FILE = 'dairy_data.xlsx'

def _get_or_create_workbook():
    if os.path.exists(EXCEL_FILE):
        return openpyxl.load_workbook(EXCEL_FILE)
    wb = openpyxl.Workbook()
    # Users sheet
    ws1 = wb.active; ws1.title = 'Users'
    ws1.append(['#','Name','Username','Email','Phone','Registered At'])
    # Orders sheet
    ws2 = wb.create_sheet('Orders')
    ws2.append(['Order#','Customer','Username','Address','Items','Total (₹)','Date'])
    _style_header(ws1); _style_header(ws2)
    return wb

def _style_header(ws):
    green = PatternFill('solid', fgColor='2e9e4f')
    bold  = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = green; cell.font = bold
        cell.alignment = Alignment(horizontal='center')

def _append_user_to_excel(name, username, email, phone):
    try:
        wb = _get_or_create_workbook()
        ws = wb['Users']
        row_num = ws.max_row
        ws.append([row_num, name, username, email, phone, datetime.now().strftime('%Y-%m-%d %H:%M')])
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f'Excel user write error: {e}')

def _append_order_to_excel(oid, name, username, address, total, cart):
    try:
        wb  = _get_or_create_workbook()
        ws  = wb['Orders']
        items_str = ', '.join(f"{v['name']} x{v['qty']}" for v in cart.values())
        ws.append([oid, name, username, address, items_str, total, datetime.now().strftime('%Y-%m-%d %H:%M')])
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f'Excel order write error: {e}')

@app.route('/export/excel')
def export_excel():
    if not session.get('is_admin'):
        flash('Admin only.', 'danger')
        return redirect(url_for('dashboard'))
    if not os.path.exists(EXCEL_FILE):
        flash('No data exported yet.', 'info')
        return redirect(url_for('dashboard'))
    return send_file(EXCEL_FILE, as_attachment=True, download_name='KavisDairy_Data.xlsx')

# ── Run ───────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=8080)
